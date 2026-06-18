# tools/write_momo.py
import json
import os
import openpyxl
from datetime import datetime, date
from openpyxl.utils import get_column_letter, column_index_from_string
from langchain.tools import tool
from typing import Optional

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "DAILY BALANCE AUGUST 2024.xlsx")
SHEET_NAME = "MOMO,AIRTEL,CARDS"

DATE_ROW = 3
FIRST_DATA_COL = "ANR"  # Earliest visible column from screenshot

PAYMENT_ROWS = {
    "MOMOPAY":   4,
    "AIRTEL":    5,
    "VISA CARD": 6,
    "RUBIS CARD":7,
    "RUBIS APP": 8,
}

def parse_excel_date(cell_value) -> Optional[date]:
    if not cell_value:
        return None
    try:
        raw = str(cell_value).replace(" ", "").replace("-", "")
        for suffix in ["st", "nd", "rd", "th"]:
            raw = raw.replace(suffix, "")
        return datetime.strptime(raw, "%d%B%Y").date()
    except ValueError:
        return None

def find_or_create_date_column(ws, target_date: date) -> str:
    start_col = column_index_from_string(FIRST_DATA_COL)

    # Search for existing column
    for col in range(start_col, ws.max_column + 1):
        if parse_excel_date(ws.cell(row=DATE_ROW, column=col).value) == target_date:
            return get_column_letter(col)

    # Not found — create after last date column
    last_used_col = start_col
    for col in range(start_col, ws.max_column + 2):
        if ws.cell(row=DATE_ROW, column=col).value and \
           parse_excel_date(ws.cell(row=DATE_ROW, column=col).value):
            last_used_col = col

    new_col = last_used_col + 1
    new_col_letter = get_column_letter(new_col)

    day = target_date.day
    suffix = "th" if 11 <= day <= 13 else {1:"st", 2:"nd", 3:"rd"}.get(day % 10, "th")
    ws.cell(row=DATE_ROW, column=new_col).value = \
        f"{day}{suffix} -{target_date.strftime('%B%Y')}"

    return new_col_letter


@tool
def write_electronic_sales_sheet(report_json: str) -> str:
    """
    Write electronic sales (MomoPay, Airtel, Visa Card, Rubis Card, Rubis App)
    into the MOMO,AIRTEL,CARDS sheet of the petrol station Excel file.
    """
    try:
        data = json.loads(report_json) if isinstance(report_json, str) else report_json

        target_date = datetime.strptime(data["date"], "%Y-%m-%d").date()

        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["MOMO,AIRTEL,CARDS"]

        col = find_or_create_date_column(ws, target_date)
        col_idx = column_index_from_string(col)

        written = []
        skipped = []

        for payment_type, value in data["electronic_sales"].items():
            payment_key = payment_type.upper().strip()

            if payment_key not in PAYMENT_ROWS:
                skipped.append(payment_key)
                continue

            if value is None:
                skipped.append(payment_key)
                continue

            row = PAYMENT_ROWS[payment_key]
            ws.cell(row=row, column=col_idx).value = value
            written.append(payment_key)

        wb.save(EXCEL_PATH)

        return json.dumps({
            "status": "success",
            "date": str(target_date),
            "column": col,
            "written": written,
            "skipped": skipped
        })

    except Exception as e:
        return json.dumps({
            "status": "error",
            "message": str(e)
        })
    
