import json
import os
from datetime import datetime, date
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from langchain_core.tools import tool

PUMP_ROWS = {
    # PMS pumps (rows 4-15)
    "PMS 1": {"opening": 4,  "closing": 5,  "sales": 6},
    "PMS 2": {"opening": 7,  "closing": 8,  "sales": 9},
    "PMS 3": {"opening": 10, "closing": 11, "sales": 12},
    "PMS 4": {"opening": 13, "closing": 14, "sales": 15},

    # AGO pumps (rows 23-31)
    "AGO 2": {"opening": 23, "closing": 24, "sales": 25},
    "AGO 3": {"opening": 26, "closing": 27, "sales": 28},
    "AGO 4": {"opening": 29, "closing": 30, "sales": 31},
}

# Summary rows per fuel type
PMS_SUMMARY = {"total": 16, "rtt": 17, "net": 18}
AGO_SUMMARY = {"total": 32, "rtt": 33, "net": 34}
DATE_ROW = 3
FIRST_DATA_COL = "ANZ"  # Earliest date column from your screenshot

@tool
def write_fuel_meter_sheet(report_json: str) -> str:
    """
    Write extracted fuel report data into the petrol station Excel file.
    Takes the JSON output from extract_fuel_report and writes each pump's
    opening, closing, and RTT into the correct row and date column.
    """

    EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "DAILY BALANCE AUGUST 2024.xlsx")

   

    

    # ── Helper: parse date header from cell e.g. "30th -May2026" ──
    def parse_excel_date(cell_value) -> Optional[date]:
        if not cell_value:
            return None
        try:
            raw = str(cell_value).replace(" ", "").replace("-", "")
            for suffix in ["st", "nd", "rd", "th"]:
                raw = raw.replace(suffix, "")
            return datetime.strptime(raw, "%d%B%Y").date()
        except ValueError:
            return None

    # ── Helper: find column letter for a given date ──
    def find_or_create_date_column(ws, target_date: date) -> str:
        start_col = column_index_from_string(FIRST_DATA_COL)

        # Search for existing column
        for col in range(start_col, ws.max_column + 1):
            cell = ws.cell(row=DATE_ROW, column=col)
            if parse_excel_date(cell.value) == target_date:
                return get_column_letter(col)

        # Not found — create new column after the last date column
        last_used_col = start_col
        for col in range(start_col, ws.max_column + 2):
            cell = ws.cell(row=DATE_ROW, column=col)
            if cell.value and parse_excel_date(cell.value):
                last_used_col = col

        new_col = last_used_col + 1
        new_col_letter = get_column_letter(new_col)

        day = target_date.day
        suffix = "th" if 11 <= day <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
        ws.cell(row=DATE_ROW, column=new_col).value = \
            f"{day}{suffix} -{target_date.strftime('%B%Y')}"

        return new_col_letter

    # ── Main logic ──
    try:
        # Accept both raw dict and JSON string
        data = json.loads(report_json) if isinstance(report_json, str) else report_json

        # Parse date
        target_date = datetime.strptime(data["date"], "%Y-%m-%d").date()

        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["METER"]

        col = find_or_create_date_column(ws, target_date)
        col_idx = column_index_from_string(col)

        written = []
        skipped = []

        # Write each pump
        for pump_name, readings in data["pumps"].items():
            if readings["opening"] is None or readings["closing"] is None:
                skipped.append(pump_name)
                continue

            rows = PUMP_ROWS[pump_name]

            # Opening and closing
            ws.cell(row=rows["opening"], column=col_idx).value = readings["opening"]
            ws.cell(row=rows["closing"], column=col_idx).value = readings["closing"]

            # Sales as Excel formula
            ws.cell(row=rows["sales"], column=col_idx).value = \
                f"={col}{rows['closing']}-{col}{rows['opening']}"

            written.append(pump_name)

        # Write PMS summary rows
        pms_rtt = data["rtt"]["PMS"] or 0
        ws.cell(row=PMS_SUMMARY["total"], column=col_idx).value = \
            f"={col}6+{col}9+{col}12+{col}15"
        ws.cell(row=PMS_SUMMARY["rtt"],   column=col_idx).value = pms_rtt
        ws.cell(row=PMS_SUMMARY["net"],   column=col_idx).value = \
            f"={col}{PMS_SUMMARY['total']}-{col}{PMS_SUMMARY['rtt']}"

        # Write AGO summary rows
        ago_rtt = data["rtt"]["AGO"] or 0
        ws.cell(row=AGO_SUMMARY["total"], column=col_idx).value = \
            f"={col}25+{col}28+{col}31"
        ws.cell(row=AGO_SUMMARY["rtt"],   column=col_idx).value = ago_rtt
        ws.cell(row=AGO_SUMMARY["net"],   column=col_idx).value = \
            f"={col}{AGO_SUMMARY['total']}-{col}{AGO_SUMMARY['rtt']}"

        wb.save(EXCEL_PATH)

        # Return summary for the agent to report back to user
        return json.dumps({
            "status": "success",
            "date": str(target_date),
            "column": col,
            "written": written,
            "skipped": skipped,
            "pms_rtt": pms_rtt,
            "ago_rtt": ago_rtt
        })

    except Exception as e:
        return json.dumps({
            "status": "error",
            "message": str(e)
        })
    


test_data = {
    "date": "2026-05-31",
    "pumps": {
        "PMS 1": {"opening": 1692332.200, "closing": 1693160.057},
        "PMS 2": {"opening": 672625.908, "closing": 672967.199},
        "PMS 3": {"opening": 1383791.259, "closing": 1384340.132},
        "PMS 4": {"opening": 345888.500, "closing": 345996.666},
        "AGO 2": {"opening": 544624.767, "closing": 544827.369},
        "AGO 3": {"opening": 444987.695, "closing": 445182.641},
        "AGO 4": {"opening": 292017.340, "closing": 292260.264}
    },
    "rtt": {"PMS": 0, "AGO": 0}
}

